import os
import shutil
import subprocess
import docx
import openpyxl
import PyPDF2 
import warnings

class search_word:

    find_file = []
    spisok_file = []

    #path indication feature
    def search(self, disk, format_files, word):

        #additional parameters
        exc = []
        doc = ('.doc', '.docx')
        exl = ('.xls', '.xlsx')
        t = 'txt'
        d = 'doc'
        e = 'exl'        
        p = 'pdf'
        a = 'all'

        #search fails with format for txt
        if format_files == t or format_files == a: 
            for adress, dirs, files in os.walk(disk):
                for file in files:
                    s = (os.path.join(adress, file))
                    if file.endswith('.txt') and '$' not in s:
                        self.spisok_file.append(s)
            #search for a word
            for x in self.spisok_file: 
                with open(x) as r:
                    try:
                        for line in r:
                            if word in line:   
                                self.find_file.append(x)  
                                break
                    except Exception as fail:
                        exc.append(r) 

        #search fails with format for docx
        if format_files == d or format_files == a: 
            for adress, dirs, files in os.walk(disk):
                for file in files:
                    s = (os.path.join(adress, file))
                    if file.endswith(doc) and '$' not in s:
                        self.spisok_file.append(s)     
            #search for a word
            for x in self.spisok_file: 
                with open(x) as r:
                    try:
                        doc = docx.Document(x)
                        text = []
                        for paragraph in doc.paragraphs:
                            text.append(paragraph.text)    
                        for line in text:
                            if word in line:                        
                                self.find_file.append(x)
                                break
                    except Exception as fail:
                        exc.append(r)   

        #search fails with format for exel
        if format_files == e or format_files == a: 
            for adress, dirs, files in os.walk(disk):
                for file in files:
                    s = (os.path.join(adress, file))
                    if file.endswith(exl) and '$' not in s:
                        self.spisok_file.append(s)
            #search for a word
            for x in self.spisok_file: 
                with open(x) as r:
                    try:
                        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                        path = x
                        wb_obj = openpyxl.load_workbook(path)
                        sheet_obj = wb_obj.active
                        max_col = sheet_obj.max_column
                        for i in range(1, max_col + 1):
                            cell_obj = sheet_obj.cell(row = 1, column = i)
                            if word in cell_obj.value:                            
                                self.find_file.append(x)
                                break          
                    except Exception as fail:
                        exc.append(r) 

        #search fails with format for pdf
        if format_files == p or format_files == a: 
            for adress, dirs, files in os.walk(disk):
                for file in files:
                    s = (os.path.join(adress, file))
                    if file.endswith('.pdf') and '$' not in s:
                        self.spisok_file.append(s)
            #search for a word
            for x in self.spisok_file: 
                with open(x) as r:
                    try:
                        warnings.filterwarnings('ignore', category=UserWarning, module='PyPDF2')
                        pdf_file = open(x, 'rb')
                        read_pdf = PyPDF2.PdfFileReader(pdf_file)
                        page = read_pdf.getPage(0)
                        page_content = page.extractText()
                        if word in page_content:                            
                            self.find_file.append(x)
                            break
                        pdf_file.close()
                    except Exception as fail:
                        exc.append(r) 

    #displaying a list of found files
    def path_file(self):
        if len(self.find_file) == 0:
            print('We did not find the words in files')
        else:
            print(",\n".join(map(str, self.find_file)))

    #list of found files
    def list(self):
        if len(self.find_file) == 0:
            print('We did not find the words in files')
        else:
            return self.find_file

    #file opening function
    def opening(self):    
        if len(self.find_file) == 0:
            print('We did not find the words in files')
        else:
            for x in self.find_file:
                head, tail = os.path.split(x)
                print('We faind faile:', tail)
                want_open = input('Open file? \n (y/n): ')
                if want_open == 'y':
                    subprocess.Popen('explorer ' + x)
        
    #copying found files
    def copying(self, path_copy = '', numcopy = 'a'):
        if len(self.find_file) == 0:
            print('We did not find the words in files')
        else:    
            try:
                os.mkdir(path_copy)
                for x in self.find_file:
                    head, tail = os.path.split(x) 
                    if numcopy != 'a':
                        print('Copy file', tail, '?' )
                        want_copy = input('y/n:')
                        if want_copy == 'y':
                            shutil.copy(x,path_copy)
                            print('File:', tail, 'was copied')
                    elif numcopy == 'a':
                        shutil.copy(x,path_copy)
                        print('File:', tail, 'was copied')
            except Exception as fail:
                print('A directory with this name already exists') 
        