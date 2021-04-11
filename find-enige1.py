import os
import subprocess
import docx
import openpyxl
import PyPDF2


class search_word:

    spisok_file1 = []
    spisok_file2 = []
    spisok_file3 = []
    spisok_file4 = []
    
    find_file1 = []
    find_file2 = []
    find_file3 = []
    find_file4 = []

    exc = []

    def search(self, disk, format_files, word):
        doc = ('.doc', '.docx')
        exl = ('.xls', '.xlsx')
        t = 'txt'
        d = 'doc'
        e = 'exl'        
        p = 'pdf'
        a = 'all'
        

        #1 <=> search fails with format for txt
        if format_files == t or format_files == a: 
            for adress, dirs, files in os.walk(disk):
                for file in files:
                    s = (os.path.join(adress, file))
                    if file.endswith('.txt') and '$' not in s:
                        self.spisok_file1.append(s)
            #2 <=> search for a word
            for x in self.spisok_file1: 
                with open(x) as r:
                    try:
                        for line in r:
                            if word in line:   
                                self.find_file1.append(x)  
                                break
                    #3 <=> adding unopened files to exceptions
                    except Exception as fail:
                        self.exc.append(r) 
            #4 <=> open file 
            if len(self.find_file1) == 0:
                print('We did not find the words in the format txt')
            else:
                for x in self.find_file1:
                    head, tail = os.path.split(x)
                    print('We faind faile:', tail)
                    want_open = input('Open file? \n (y/n): ')
                    if want_open == 'y':
                        subprocess.Popen('explorer ' + x)
                
      
        #1 for docx
        if format_files == d or format_files == a: 
            for adress, dirs, files in os.walk(disk):
                for file in files:
                    s = (os.path.join(adress, file))
                    if file.endswith(doc) and '$' not in s:
                        self.spisok_file2.append(s)
            #2
            for x in self.spisok_file2: 
                with open(x) as r:
                    try:
                        doc = docx.Document(x)
                        text = []
                        for paragraph in doc.paragraphs:
                            text.append(paragraph.text)    
                        for line in text:
                            if word in line:                        
                                self.find_file2.append(x)
                                break
                    #3
                    except Exception as fail:
                        self.exc.append(r)
            #4   
            if len(self.find_file2) == 0:
                print('We did not find the words in the format doc')
            else:
                for x in self.find_file2:
                    head, tail = os.path.split(x)
                    print('We faind faile:', tail)
                    want_open = input('Open file? \n (y/n): ')
                    if want_open == 'y':
                        subprocess.Popen('explorer ' + x)


        #1 for exel
        if format_files == e or format_files == a: 
            for adress, dirs, files in os.walk(disk):
                for file in files:
                    s = (os.path.join(adress, file))
                    if file.endswith(exl) and '$' not in s:
                        self.spisok_file3.append(s)
            #2
            for x in self.spisok_file3: 
                with open(x) as r:
                    try:
                        path = x
                        wb_obj = openpyxl.load_workbook(path)
                        sheet_obj = wb_obj.active
                        max_col = sheet_obj.max_column
                        for i in range(1, max_col + 1):
                            cell_obj = sheet_obj.cell(row = 1, column = i)
                            if word in cell_obj.value:                            
                                self.find_file3.append(x)
                                break
                    #3                                
                    except Exception as fail:
                        self.exc.append(r) 
            #4
            if len(self.find_file3) == 0:
                print('We did not find the words in the format exl')
            else:
                for x in self.find_file3:
                    head, tail = os.path.split(x)
                    print('We faind faile:', tail)
                    want_open = input('Open file? \n (y/n): ')
                    if want_open == 'y':
                        subprocess.Popen('explorer ' + x)


        #1 for pdf
        if format_files == p or format_files == a: 
            for adress, dirs, files in os.walk(disk):
                for file in files:
                    s = (os.path.join(adress, file))
                    if file.endswith('.pdf') and '$' not in s:
                        self.spisok_file4.append(s)
            #2
            for x in self.spisok_file4: 
                with open(x) as r:
                    try:
                        pdf_file = open(x, 'rb')
                        read_pdf = PyPDF2.PdfFileReader(pdf_file)
                        page = read_pdf.getPage(0)
                        page_content = page.extractText()
                        if word in page_content:                            
                            self.find_file4.append(x)
                            break
                        pdf_file.close()
                    #3
                    except Exception as fail:
                        self.exc.append(r) 
            #4
            if len(self.find_file4) == 0:
                print('We did not find the words in the format pdf')
            else:
                for x in self.find_file4:
                    head, tail = os.path.split(x)
                    print('We faind faile:', tail)
                    want_open = input('Open file? \n (y/n): ')
                    if want_open == 'y':
                        subprocess.Popen('explorer ' + x)